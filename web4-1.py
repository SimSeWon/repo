from flask import Flask, render_template, request
app=Flask(__name__)

import requests
from bs4 import BeautifulSoup

#엑셀 import
from openpyxl import Workbook
write_wb = Workbook()
write_ws = write_wb.active

# write_ws.cell(1,1,"안녕")
#
# write_ws.cell(2,1,"여긴느 2,1")
#
# write_wb.save(1,2,'test.xlsx')

@app.route('/')
def hello():
    return render_template("index.html")

@app.route('/result', methods=['POST'])
def hello2():

    result_list=[]

    keyword = request.form['input1']
    page = request.form['input2']

    print(keyword)
    print(page)

    for i in range(1, int(page)+1):
        req = requests.get('https://search.daum.net/search?w=news&nil_search=btn&DA=NTB&enc=utf8&cluster=y&cluster_page'+str(i)+'&q=' + keyword)
        soup = BeautifulSoup(req.text, 'html.parser')

    for i in soup.find_all('a', class_='f_link_b'):
        print(i.text)


    #https://search.daum.net/search?w=tot&DA=YZR&t__nil_searchbox=btn&sug=&sugo=&q=%ED%95%9C%EA%B5%AD
    req=requests.get('https://search.daum.net/search?w=tot&DA=YZR&t__nil_searchbox=btn&sug=&sugo=&q='+keyword)
    soup=BeautifulSoup(req.text,'html.parser')

    print(soup.select('#clusterResultUL > li'))

    #제목뽑기
    #키워드, 페이지수대로 대이터뽑기
    #엑셀파일 저장

    for i in soup.find_all('a', class_='f_link_b'):

        result_list.append(i.text)

    for i in range(1,len(result_list)+1):
        print(result_list[i-1])
        write_ws.cell(i,1,result_list[i-1])

    write_wb.save('resu66lt11.xlsx')
     #0리스트는 0부터 시작하기때문

    return render_template('result.html', result=result_list)


if __name__ == '__main__':
    app.run()